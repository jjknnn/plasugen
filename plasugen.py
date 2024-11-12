import random
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill


def open_file():
    """
    Read xlsx file
    """
    names = {}
    wb = load_workbook(filename="nimilista.xlsx")

    sheet = wb.active

    # Fetch names from xlsx based on columns
    for count, column in enumerate(sheet.iter_cols(values_only=True)):
        none_removed_list = []
        # Remove empty cells
        for val in list(column):
            if val is not None:
                none_removed_list.append(val)
        names[count] = none_removed_list
    return names


def ask_user_tables(count):
    """
    Ask user for table sizes and amounts
    """
    ans = True
    i = 0
    print("\nMääritetään pöytien määrät ja koot")
    tables = {}
    while ans:
        print(f"Nimiä jäljellä: {count}")
        places = input("Syötä pöydän paikkamäärä: ")
        amount = input("Syötä kyseisten pöytien määrä: ")
        if int(places) * int(amount) > count:
            print(
                "\nSyötit enemmän paikkoja kun listassa on nimiä!\nPöytää ei lisätty!"
            )
            continue
        a = 0
        while a < int(amount):
            tables[i] = places
            a += 1
            i += 1
        count -= int(places) * int(amount)
        if count == 0:
            print("Kaikki paikat täytetty!")
            break
        i += 1
        cont = input("Jatketaanko? y/n ")
        if cont != "y":
            ans = False
        print()  # Newline to help readability
    return tables


def ask_user_type():
    """
    Ask user for name sorting type
    """
    print("Miten haluat pöydät järjestettävän: ")
    while True:
        inp = input("R(andom) / J(ärjestys) /K(averi): ")
        if inp.lower() in ["r", "k", "j"]:
            break
        else:
            print("Hyväksytyt syötteet:  r, j, tai k.")
    return inp[0]


def fill_names_random(names, tables):
    """
    Fill names to given tables randomly
    """
    tables_with_names = {}
    all_names = []
    for name in names:
        all_names.extend(names[name])

    # Shuffle names
    random.shuffle(all_names)

    # Generate tables with names
    for i, table in enumerate(tables):
        tables_with_names[i] = make_tuples(all_names[: int(tables[table])])
        all_names = all_names[int(tables[table]) :]
    return tables_with_names


def make_tuples(input_list):
    """
    Makes tuples from the names for feeding to tables
    """
    result = []
    if len(input_list) % 2 == 0:
        for i in range(0, len(input_list), 2):
            result.append((input_list[i], input_list[i + 1]))
    else:
        for i in range(0, len(input_list) - 1, 2):
            result.append((input_list[i], input_list[i + 1]))
        result.append((input_list[i + 2], ""))
    return result


def make_file(generated, colors_for_names):
    """
    Make xlsx file with A and B column with names
    """
    wb = Workbook()
    ws = wb.active
    mem = 1
    for names in generated:
        for name in generated[names]:
            if len(name) >= 2:
                ws[f"A{mem}"], ws[f"B{mem}"] = name
                mem += 1
            elif len(name) == 1:
                ws[ws[f"A{mem}"]] = name
            else:
                continue
        ws[f"A{mem}"], ws[f"B{mem}"] = ("", "")
        mem += 1
    if colors_for_names:
        ws = color_workbook(ws, colors_for_names)
    wb.save("names.xlsx")


def color_workbook(ws, colors_for_names):
    """
    Loop through all cells and apply colors based on the value in given color dictionary
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in colors_for_names:
                # Get the RGB color from the dictionary and convert it to hex
                rgb_color = colors_for_names[cell.value]
                hex_color = rgb_to_hex(rgb_color)

                # Apply the background color
                fill = PatternFill(
                    start_color=hex_color, end_color=hex_color, fill_type="solid"
                )
                cell.fill = fill
    return ws


def rgb_to_hex(rgb):
    """Helper function to return hex from rgb"""
    return "{:02X}{:02X}{:02X}".format(*rgb)


def balance_table(names):
    """
    Generate names to such order that they are evenly divided to the tables
    TODO: Reformat this method. It's a disaster :)
    """
    completed_name_list = []
    list_list = []
    for list in names:
        list_list.append(names[list])
    for j, name_list in enumerate(list_list, 1):
        if j < len(list_list):
            if len(list_list[j]) < len(list_list[j - 1]):
                list1 = list_list[j - 1]
                list2 = list_list[j]
            else:
                list2 = list_list[j - 1]
                list1 = list_list[j]
        else:
            break
        len1 = len(list1)
        len2 = len(list2)
        total_length = len1 + len2
        repetitions = total_length // len2
        merged_list = []
        for i in range(max(len1, len2)):
            if i < len1:
                merged_list.append(list1[i])
            if i % repetitions == 0 and i // repetitions < len2:
                merged_list.append(list2[i // repetitions])
        list_list[j] = merged_list
        completed_name_list = merged_list
    no_dups = []
    for item in completed_name_list:
        if item not in no_dups:
            no_dups.append(item)
    for name in names:
        for n in names[name]:
            if n not in no_dups:
                random_index = random.randint(0, len(no_dups))
                no_dups.insert(random_index, n)
    return no_dups


def fill_names_order(names, tables):
    """
    Fill names to tables when using ordered method
    """
    balanced = balance_table(names)
    tables_with_names = {}
    for i, table in enumerate(tables):
        tables_with_names[i] = make_tuples(balanced[: int(tables[table])])
        balanced = balanced[int(tables[table]) :]
    return tables_with_names


def fill_names_buddies(names, tables, groups):
    """
    Documentation wow
    """
    group_sizes = get_group_sizes(names)

    # sorted_items = sorted(group_sizes.items(), key=lambda x: x[1], reverse=True)
    table_sizes = [int(tables[table]) for table in tables]
    print(
        "Kaverisitsien paikkajako kestää pitkään jos monta epämääräistä pöytää. \
          \nOdotathan hetken."
    )
    while True:
        shuffled_groups = shuffle_groups(group_sizes)
        not_all_filled, seat_distribution = try_to_sort_to_tables(
            shuffled_groups, table_sizes
        )
        if not_all_filled:
            continue
        else:
            break
    print("Löytyi")
    tables_with_names = add_names(seat_distribution, groups)
    return tables_with_names


def add_names(distribution, groups):
    """
    Fills group member names from the given groups dictionary
    """
    tables = {}
    for i, table in enumerate(distribution):
        table_with_all_names = []
        for name in distribution[table][2]:
            table_with_all_names.extend(groups[name])
        tables[i] = make_tuples(table_with_all_names)
    return tables


def try_to_sort_to_tables(groups, tables):
    """
    Tries to force groups into tables
    If perfect match returns False as not_added
    """
    not_added = False
    tables_with_names = {}
    for i, table in enumerate(tables, 0):
        tables_with_names[i] = [table, 0, []]

    for j, (name, amount) in enumerate(groups, 0):
        n = j % (i + 1)
        if tables_with_names[n][1] <= tables_with_names[n][1] + amount:
            tables_with_names[n][1] += amount  # Add the amount to table
            tables_with_names[n][2].append(name)  # Add the name(s) to the list of names
        if tables_with_names[n][1] > tables_with_names[n][0]:
            not_added = True
            continue
    return not_added, tables_with_names


def get_group_sizes(names):
    """
    Returns how many occurrances of a name is in given list as a dictionary
    """
    # Remove dictionary structure used to identify different subgroups
    # (f.ex different organization groups etc)
    names = names[0]
    group_sizes = {}
    for name in names:
        if not group_sizes.get(name):
            group_sizes[name] = 1
        else:
            group_sizes[name] += 1
    return group_sizes


def shuffle_groups(groups):
    """
    Helper function to shuffle group order
    """
    groups_as_tuple = list(groups.items())
    random.shuffle(groups_as_tuple)
    return tuple(groups_as_tuple)


def make_file_order(generated):
    """
    Makes the xlsx file when using ordered method
    """
    wb = Workbook()
    ws = wb.active
    mem = 1
    for names in generated:
        for name in generated[names]:
            ws[f"A{mem}"], ws[f"B{mem}"] = name
            mem += 1
        ws[f"A{mem}"], ws[f"B{mem}"] = ("", "")
        mem += 1
    wb.save("names.xlsx")


def shuffle_names(names):
    """
    Shuffles names from xlsx
    """
    shuffled = {}
    for name in names:
        if len(names[name]) != 0:
            random.shuffle(names[name])
            shuffled[name] = names[name]
    return shuffled


def parse_buddy_names_from_file():
    """
    Get the names of buddies from the .xlsx file
    Returns groups and colours for groups
    """
    color_dictionary = {}
    wb = load_workbook(filename="nimilista.xlsx")
    sheet = wb.active
    buddy_dictionary = {}
    # Fetch names from xlsx based on columns
    for column in sheet.iter_rows(values_only=True):
        group_color = generate_group_color()
        none_removed_list = []
        # Remove empty cells
        for val in list(column):
            if val is not None:
                none_removed_list.append(val)
        buddy_dictionary[column[0]] = none_removed_list
        for name in none_removed_list:
            color_dictionary[name] = group_color
    return buddy_dictionary, color_dictionary


def generate_group_color():
    """
    Generates a random color in rgb
    """
    r = random.randint(0, 255)
    g = random.randint(0, 255)
    b = random.randint(0, 255)
    return (r, g, b)


def duplicate_names_to_match_group_size(names, groups):
    """
    Duplicates a name to make matching to tables easier
    """
    all_names = []
    names = {}
    for group in groups:
        for _ in range(len(groups[group])):
            all_names.append(group)
    names[0] = all_names
    return names


def main():
    names = open_file()
    organization_type = ask_user_type()
    if organization_type.lower() == "k":
        names_with_buddy_names, color_dictionary = parse_buddy_names_from_file()
        names = duplicate_names_to_match_group_size(names, names_with_buddy_names)
    else:
        color_dictionary = False
    amount = 0
    for name in names:
        amount += len(names[name])
    tables = ask_user_tables(amount)
    if organization_type.lower() == "r":
        names = shuffle_names(names)
        generated = fill_names_random(names, tables)
    elif organization_type.lower() == "j":
        names = shuffle_names(names)
        generated = fill_names_order(names, tables)
    elif organization_type.lower() == "k":
        generated = fill_names_buddies(names, tables, names_with_buddy_names)
    make_file(generated, color_dictionary)
    print("names.xlsx luotu! Katso kansion sisältö!")


main()
